import os
import csv
import openpyxl

# PROGRAM TO GO THROUGH EACH FILE IN FOLDER AND PRINT/WRITE MATERIAL & OVERALL CELLS IN TAKEOFF
# Declaring locations
directory = '/home/abramnaranjo/.bkp/knells/FIRSTFLOOR'
folder = '/home/abramnaranjo/.bkp/knells/land'
writetofile = '/home/abramnaranjo/.bkp/FIRSTFLOOR/output.csv'

# For each file in folder $path
for filename in os.listdir(directory):
    path = os.path.join(directory, filename)
    wb_obj = openpyxl.load_workbook(path)
    # Makes "Setup" the active sheet in takeoff program
    sh = wb_obj['Setup']
    # Gets the project name from cell B1 in the Setup excel spreadsheet
    project_name = str(sh.cell(row=1, column=2).value)
    # Changes the active spreadsheet to "Overall"
    sh = wb_obj['Overall']
    # Gets cell range A5:F7, rounded if it is floating integer, and project_total cell F8
    #project_total = []
    cellA5 = sh.cell(row=5, column=1).value
    cellB5 = sh.cell(row=5, column=2).value
    cellC5 = sh.cell(row=5, column=3).value
    cellD5 = sh.cell(row=5, column=4).value
    cellE5 = sh.cell(row=5, column=5).value
    cellF5 = sh.cell(row=5, column=6).value
    cellA6 = sh.cell(row=6, column=1).value
    cellB6 = round(sh.cell(row=6, column=2).value)
    cellC6 = round(sh.cell(row=6, column=3).value)
    cellD6 = round(sh.cell(row=6, column=4).value)
    cellE6 = round(sh.cell(row=6, column=5).value)
    cellF6 = round(sh.cell(row=6, column=6).value)
    cellA7 = sh.cell(row=7, column=1).value
    cellB7 = round(sh.cell(row=7, column=2).value)
    cellC7 = round(sh.cell(row=7, column=3).value)
    cellD7 = round(sh.cell(row=7, column=4).value)
    cellE7 = round(sh.cell(row=7, column=5).value)
    cellF7 = round(sh.cell(row=7, column=6).value)
    #cellF8Total = int(round(sh.cell(row=8, column=6).value))
    #project_total.extend(cellF8Total)

    # Prints data to console
    print("")
    print(project_name)
    print(cellA5, cellB5, cellC5, cellD5, cellE5, cellF5)
    print(cellA6, cellB6, cellC6, cellD6, cellE6, cellF6)
    print(cellA7, cellB7, cellC7, cellD7, cellE7, cellF7, project_total)
    print("")

    # Declaring variables to write lines to csv files
    line0write = project_name
    line1write = cellA5, cellB5, cellC5, cellD5, cellE5, cellF5
    line2write = cellA6, cellB6, cellC6, cellD6, cellE6, cellF6
    line3write = cellA7, cellB7, cellC7, cellD7, cellE7, cellF7, project_total
    line4write = ""

    # Changes active spreadsheet to "Material" to print out all materials in takeoff, written after the "Overall" cell range above
    sh = wb_obj['Material']
    cellA1 = sh.cell(row=1, column=1).value
    cellA2 = sh.cell(row=2, column=1).value
    cellA3 = sh.cell(row=3, column=1).value
    cellA4 = sh.cell(row=4, column=1).value
    cellA5 = sh.cell(row=5, column=1).value
    cellA6 = sh.cell(row=6, column=1).value
    cellA7 = sh.cell(row=7, column=1).value
    cellA8 = sh.cell(row=8, column=1).value
    cellA9 = sh.cell(row=9, column=1).value
    cellB1 = sh.cell(row=1, column=2).value
    cellB2 = sh.cell(row=2, column=2).value
    cellB3 = sh.cell(row=3, column=2).value
    cellB4 = sh.cell(row=4, column=2).value
    cellB5 = sh.cell(row=5, column=2).value
    cellB6 = sh.cell(row=6, column=2).value
    cellB7 = sh.cell(row=7, column=2).value
    cellB8 = sh.cell(row=8, column=2).value
    cellB9 = sh.cell(row=9, column=2).value
    cellC1 = sh.cell(row=1, column=3).value
    cellC2 = sh.cell(row=2, column=3).value
    cellC3 = sh.cell(row=3, column=3).value
    cellC4 = sh.cell(row=4, column=3).value
    cellC5 = sh.cell(row=5, column=3).value
    cellC6 = sh.cell(row=6, column=3).value
    cellC7 = sh.cell(row=7, column=3).value
    cellC8 = sh.cell(row=8, column=3).value
    cellC9 = sh.cell(row=9, column=3).value
    cellD1 = sh.cell(row=1, column=4).value
    cellD2 = sh.cell(row=2, column=4).value
    cellD3 = sh.cell(row=3, column=4).value
    cellD4 = sh.cell(row=4, column=4).value
    cellD5 = sh.cell(row=5, column=4).value
    cellD6 = sh.cell(row=6, column=4).value
    cellD7 = sh.cell(row=7, column=4).value
    cellD8 = sh.cell(row=8, column=4).value
    cellD9 = sh.cell(row=9, column=4).value
    cellE1 = sh.cell(row=1, column=5).value
    cellE2 = sh.cell(row=2, column=5).value
    cellE3 = sh.cell(row=3, column=5).value
    cellE4 = sh.cell(row=4, column=5).value
    cellE5 = sh.cell(row=5, column=5).value
    cellE6 = sh.cell(row=6, column=5).value
    cellE7 = sh.cell(row=7, column=5).value
    cellE8 = sh.cell(row=8, column=5).value
    cellE9 = sh.cell(row=9, column=5).value
    cellF1 = sh.cell(row=1, column=6).value
    cellF2 = sh.cell(row=2, column=6).value
    cellF3 = sh.cell(row=3, column=6).value
    cellF4 = sh.cell(row=4, column=6).value
    cellF5 = sh.cell(row=5, column=6).value
    cellF6 = sh.cell(row=6, column=6).value
    cellF7 = sh.cell(row=7, column=6).value
    cellF8 = sh.cell(row=8, column=6).value
    cellF9 = sh.cell(row=9, column=6).value

    cells1 = cellA1, cellB1, cellC1, cellD1, cellE1, cellF1
    cells2 = cellA2, cellB2, cellC2, cellD2, cellE2, cellF2
    cells3 = cellA3, cellB3, cellC3, cellD3, cellE3, cellF3
    cells4 = cellA4, cellB4, cellC4, cellD4, cellE4, cellF4
    cells5 = cellA5, cellB5, cellC5, cellD5, cellE5, cellF5
    cells6 = cellA6, cellB6, cellC6, cellD6, cellE6, cellF6
    cells7 = cellA7, cellB7, cellC7, cellD7, cellE7, cellF7
    cells8 = cellA8, cellB8, cellC8, cellD8, cellE8, cellF8
    cells9 = cellA9, cellB9, cellC9, cellD9, cellE9, cellF9

    # Prints them to output
    print(cells1)
    print(cells2)
    print(cells3)
    print(cells4)
    print(cells5)
    print(cells6)
    print(cells7)
    print(cells8)
    print(cells9)

    # Writes out all output and lines to csv file for printing breakdown or further use in excel
    with open('out.csv', 'a') as output:
        writer = csv.writer(output)
        #writer.writerow(line0write)
        writer.writerow(line0write)
        writer.writerow(line1write)
        writer.writerow(line2write)
        writer.writerow(line3write)
        writer.writerow(line4write)
        writer.writerow(cells1)
        writer.writerow(cells2)
        writer.writerow(cells3)
        writer.writerow(cells4)
        writer.writerow(cells5)
        writer.writerow(cells6)
        writer.writerow(cells7)
        writer.writerow(cells8)
        writer.writerow(cells9)
        writer.writerow(line4write)
        output.close()
#total = sum(project_total)
#print(total)


